"""
Módulo de generación de reportes para la Calculadora de Turnos.

Este módulo maneja la generación de reportes en diferentes formatos
(Excel, PDF, correo electrónico) con los resultados de los cálculos.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Union
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import EXPORT_CONFIG, OUTPUT_DATA_DIR

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generador de reportes para los resultados de turnos."""
    
    def __init__(self):
        """Inicializa el generador de reportes."""
        self.export_config = EXPORT_CONFIG
        self.output_dir = OUTPUT_DATA_DIR
        
    def generar_reporte_completo(self, 
                                df_examenes: pd.DataFrame,
                                resultado_economico: Dict[str, any],
                                nombre_doctor: str,
                                output_path: Optional[Union[str, Path]] = None) -> Path:
        """
        Genera un reporte completo en Excel con múltiples hojas.
        
        Args:
            df_examenes: DataFrame con los exámenes procesados
            resultado_economico: Diccionario con los resultados económicos
            nombre_doctor: Nombre del doctor para el reporte
            output_path: Ruta opcional para guardar el archivo
            
        Returns:
            Path al archivo generado
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.output_dir / f"Reporte_Turnos_{timestamp}.xlsx"
        else:
            output_path = Path(output_path)
        
        # Asegurar que el directorio existe
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Crear el archivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Hoja 1: Resumen económico
            self._generar_hoja_resumen(writer, resultado_economico, nombre_doctor)
            
            # Hoja 2: Detalle de exámenes
            self._generar_hoja_examenes(writer, df_examenes, 'Todos los Exámenes')
            
            # Hoja 3: Solo TAC
            df_tac = df_examenes[df_examenes['tipo'] == 'TAC']
            if not df_tac.empty:
                self._generar_hoja_examenes(writer, df_tac, 'Exámenes TAC')
            
            # Hoja 4: Solo RX
            df_rx = df_examenes[df_examenes['tipo'] == 'RX']
            if not df_rx.empty:
                self._generar_hoja_examenes(writer, df_rx, 'Exámenes RX')
            
            # Hoja 5: Estadísticas
            self._generar_hoja_estadisticas(writer, df_examenes)
        
        # Aplicar formato adicional al archivo
        self._aplicar_formato_excel(output_path)
        
        logger.info(f"Reporte generado exitosamente: {output_path}")
        return output_path
    
    def _generar_hoja_resumen(self, writer: pd.ExcelWriter, 
                             resultado: Dict[str, any], 
                             nombre_doctor: str):
        """Genera la hoja de resumen económico."""
        # Crear DataFrame con el resumen
        data = {
            'Concepto': [],
            'Cantidad': [],
            'Tarifa Unitaria': [],
            'Total': []
        }
        
        # Agregar líneas del resumen
        conceptos = [
            ('Horas Trabajadas', resultado['horas_trabajadas'], 55000, resultado['honorarios_hora']),
            ('Exámenes RX', resultado['rx_count'], 5300, resultado['rx_total']),
            ('Exámenes TAC Simple', resultado['tac_count'], 42300, resultado['tac_total']),
            ('Exámenes TAC Doble', resultado['tac_doble_count'], 84600, resultado['tac_doble_total']),
            ('Exámenes TAC Triple', resultado['tac_triple_count'], 126900, resultado['tac_triple_total'])
        ]
        
        for concepto, cantidad, tarifa, total in conceptos:
            data['Concepto'].append(concepto)
            data['Cantidad'].append(cantidad)
            data['Tarifa Unitaria'].append(f"${tarifa:,.0f}")
            data['Total'].append(f"${total:,.0f}")
        
        # Agregar línea de total
        data['Concepto'].append('TOTAL GENERAL')
        data['Cantidad'].append('')
        data['Tarifa Unitaria'].append('')
        data['Total'].append(f"${resultado['total']:,.0f}")
        
        df_resumen = pd.DataFrame(data)
        
        # Escribir a Excel
        df_resumen.to_excel(writer, sheet_name='Resumen Económico', index=False)
        
        # Obtener la hoja para aplicar formato
        worksheet = writer.sheets['Resumen Económico']
        
        # Agregar título
        worksheet.insert_rows(1, 3)
        worksheet['A1'] = f'RESUMEN DE TURNOS - DR. {nombre_doctor.upper()}'
        worksheet['A2'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Aplicar formato al título
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A2'].font = Font(size=12, italic=True)
        worksheet.merge_cells('A1:D1')
        worksheet.merge_cells('A2:D2')
    
    def _generar_hoja_examenes(self, writer: pd.ExcelWriter, 
                              df: pd.DataFrame, 
                              nombre_hoja: str):
        """Genera una hoja con el detalle de exámenes."""
        # Seleccionar columnas relevantes
        columnas = ['fecha', 'hora', 'paciente', 'procedimiento', 'sala', 'tipo']
        columnas_disponibles = [col for col in columnas if col in df.columns]
        
        df_export = df[columnas_disponibles].copy()
        
        # Renombrar columnas para mejor presentación
        rename_dict = {
            'fecha': 'Fecha',
            'hora': 'Hora',
            'paciente': 'Paciente',
            'procedimiento': 'Procedimiento',
            'sala': 'Sala',
            'tipo': 'Tipo Examen'
        }
        df_export.rename(columns=rename_dict, inplace=True)
        
        # Escribir a Excel
        df_export.to_excel(writer, sheet_name=nombre_hoja, index=False)
    
    def _generar_hoja_estadisticas(self, writer: pd.ExcelWriter, df: pd.DataFrame):
        """Genera una hoja con estadísticas de los exámenes."""
        stats_data = []
        
        # Estadísticas por tipo de examen
        stats_data.append(['ESTADÍSTICAS POR TIPO DE EXAMEN', ''])
        stats_data.append(['Tipo', 'Cantidad'])
        
        for tipo, count in df['tipo'].value_counts().items():
            stats_data.append([tipo, count])
        
        stats_data.append(['', ''])  # Línea vacía
        
        # Estadísticas por sala
        stats_data.append(['ESTADÍSTICAS POR SALA', ''])
        stats_data.append(['Sala', 'Cantidad'])
        
        for sala, count in df['sala'].value_counts().head(10).items():
            stats_data.append([sala, count])
        
        # Crear DataFrame y escribir
        df_stats = pd.DataFrame(stats_data)
        df_stats.to_excel(writer, sheet_name='Estadísticas', 
                         index=False, header=False)
    
    def _aplicar_formato_excel(self, file_path: Path):
        """Aplica formato adicional al archivo Excel."""
        try:
            # Cargar el archivo
            wb = openpyxl.load_workbook(file_path)
            
            # Formato para cada hoja
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Aplicar bordes y formato a la primera fila (encabezados)
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D3D3D3", 
                                              end_color="D3D3D3", 
                                              fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
            
            # Guardar cambios
            wb.save(file_path)
            wb.close()
            
        except Exception as e:
            logger.warning(f"No se pudo aplicar formato adicional: {e}")
    
    def generar_correo(self, resultado_economico: Dict, nombre_doctor: str) -> str:
        """
        Genera el contenido de un correo con el resumen de honorarios.
        
        Args:
            resultado_economico: Diccionario con los resultados económicos
            nombre_doctor: Nombre del doctor
            
        Returns:
            Contenido del correo formateado
        """
        fecha_actual = datetime.now().strftime("%d de %B de %Y")
        mes_actual = datetime.now().strftime("%B %Y")
        
        correo = f"""
Asunto: Informe de Honorarios - {mes_actual} - {nombre_doctor}

Estimados,

Adjunto el detalle de honorarios correspondiente al mes de {mes_actual} para {nombre_doctor}.

RESUMEN EJECUTIVO
================

Total de horas trabajadas: {resultado_economico['horas_trabajadas']} horas
Total de exámenes realizados: {resultado_economico['rx_count'] + resultado_economico['tac_count'] + resultado_economico['tac_doble_count'] + resultado_economico['tac_triple_count']}

DESGLOSE DE HONORARIOS
=====================

1. HONORARIOS POR HORAS TRABAJADAS
   - Horas: {resultado_economico['horas_trabajadas']}
   - Tarifa por hora: $55,000
   - Subtotal: ${resultado_economico['honorarios_hora']:,.0f}

2. HONORARIOS POR EXÁMENES REALIZADOS

   a) Radiografías (RX)
      - Cantidad: {resultado_economico['rx_count']}
      - Tarifa unitaria: $5,300
      - Subtotal: ${resultado_economico['rx_total']:,.0f}

   b) TAC Simple
      - Cantidad: {resultado_economico['tac_count']}
      - Tarifa unitaria: $42,300
      - Subtotal: ${resultado_economico['tac_total']:,.0f}

   c) TAC Doble
      - Cantidad: {resultado_economico['tac_doble_count']}
      - Tarifa unitaria: $84,600
      - Subtotal: ${resultado_economico['tac_doble_total']:,.0f}

   d) TAC Triple
      - Cantidad: {resultado_economico['tac_triple_count']}
      - Tarifa unitaria: $126,900
      - Subtotal: ${resultado_economico['tac_triple_total']:,.0f}

TOTAL A PAGAR
=============
${resultado_economico['total']:,.0f}

OBSERVACIONES
============
- Los TAC dobles corresponden a estudios que incluyen dos regiones anatómicas
- Los TAC triples corresponden a estudios que incluyen tres o más regiones anatómicas
- Las tarifas aplicadas corresponden a las vigentes para el período informado

Quedo atento a cualquier consulta.

Saludos cordiales,
[Tu nombre]
[Fecha: {fecha_actual}]
"""
        
        return correo
    
    def exportar_csv(self, df: pd.DataFrame, output_path: Union[str, Path]) -> Path:
        """
        Exporta un DataFrame a formato CSV.
        
        Args:
            df: DataFrame a exportar
            output_path: Ruta de salida
            
        Returns:
            Path al archivo generado
        """
        output_path = Path(output_path)
        
        # Configuración CSV
        csv_config = self.export_config['csv']
        
        df.to_csv(
            output_path,
            index=False,
            encoding=csv_config['encoding'],
            sep=csv_config['delimiter'],
            quotechar=csv_config['quotechar']
        )
        
        logger.info(f"CSV exportado: {output_path}")
        return output_path
    
    def generar_resumen_texto(self, resultado_economico: Dict[str, any]) -> str:
        """
        Genera un resumen en texto plano de los resultados.
        
        Args:
            resultado_economico: Diccionario con los resultados
            
        Returns:
            String con el resumen
        """
        resumen = "RESUMEN DE CÁLCULO DE TURNOS\n"
        resumen += "=" * 40 + "\n\n"
        
        resumen += f"Horas trabajadas: {resultado_economico['horas_trabajadas']}\n"
        resumen += f"Total RX: {resultado_economico['rx_count']} exámenes\n"
        resumen += f"Total TAC simple: {resultado_economico['tac_count']} exámenes\n"
        
        if resultado_economico['tac_doble_count'] > 0:
            resumen += f"Total TAC doble: {resultado_economico['tac_doble_count']} exámenes\n"
        
        if resultado_economico['tac_triple_count'] > 0:
            resumen += f"Total TAC triple: {resultado_economico['tac_triple_count']} exámenes\n"
        
        resumen += "\n" + "-" * 40 + "\n"
        resumen += f"TOTAL A PAGAR: ${resultado_economico['total']:,.0f}\n"
        
        return resumen


# Funciones de utilidad para compatibilidad
def generar_reporte_excel(df_examenes: pd.DataFrame, 
                         resultado_economico: Dict[str, any],
                         nombre_doctor: str,
                         output_path: str) -> str:
    """Función de compatibilidad para generar reporte Excel."""
    generator = ReportGenerator()
    path = generator.generar_reporte_completo(
        df_examenes, resultado_economico, nombre_doctor, output_path
    )
    return str(path)


def generar_correo_resumen(resultado_economico: Dict[str, any], 
                          nombre_doctor: str) -> str:
    """Función de compatibilidad para generar correo."""
    generator = ReportGenerator()
    return generator.generar_correo(resultado_economico, nombre_doctor) 