#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para analizar un archivo Excel problemático y preprocesarlo
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook

def analizar_excel_problematico(ruta_excel):
    """Analiza un archivo Excel problemático para identificar problemas."""
    print(f"Analizando archivo: {ruta_excel}")
    
    # 1. Listar las hojas usando openpyxl (método más bajo nivel)
    try:
        workbook = load_workbook(filename=ruta_excel, read_only=True, data_only=True)
        print(f"Hojas en el archivo: {workbook.sheetnames}")
        
        # Cerrar workbook
        workbook.close()
    except Exception as e:
        print(f"Error al leer hojas con openpyxl: {e}")
    
    # 2. Intentar leer con opciones específicas para este tipo de error
    try:
        # Intentar con convert_float=False para evitar conversiones problemáticas
        df = pd.read_excel(ruta_excel, engine='openpyxl', convert_float=False)
        print(f"Lectura exitosa con convert_float=False. Dimensiones: {df.shape}")
        print("Primeras 3 filas:")
        print(df.head(3))
    except Exception as e:
        print(f"Error con convert_float=False: {e}")
    
    # 3. Intentar con otro motor de Excel
    try:
        df = pd.read_excel(ruta_excel, engine='xlrd')
        print(f"Lectura exitosa con motor xlrd. Dimensiones: {df.shape}")
    except Exception as e:
        print(f"Error con motor xlrd: {e}")
    
    # 4. Solución especial: procesar el archivo con manejo de errores personalizado
    try:
        # Cargar el archivo Excel usando openpyxl
        wb = load_workbook(filename=ruta_excel, read_only=True, data_only=True)
        
        # Obtener la primera hoja (ajustar según sea necesario)
        sheet = wb.active
        
        # Extraer datos manualmente
        data = []
        headers = []
        
        # Obtener encabezados
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # Procesar cada fila
        for row in sheet.iter_rows(min_row=2):
            row_data = []
            for cell in row:
                # Manejar específicamente el valor '0, 0'
                if cell.value == '0, 0':
                    row_data.append(0)
                else:
                    row_data.append(cell.value)
            data.append(row_data)
        
        # Crear DataFrame
        df_manual = pd.DataFrame(data, columns=headers)
        print("Lectura manual exitosa:")
        print(f"Dimensiones: {df_manual.shape}")
        print("Primeras 3 filas:")
        print(df_manual.head(3))
        
        # Guardar como CSV para procesamiento posterior
        csv_path = ruta_excel.replace('.xlsx', '_preprocesado.csv')
        df_manual.to_csv(csv_path, index=False)
        print(f"Archivo CSV guardado en: {csv_path}")
        
        # Cerrar workbook
        wb.close()
        
        return {
            "exito": True,
            "csv_path": csv_path,
            "filas": len(df_manual),
            "columnas": df_manual.columns.tolist()
        }
    except Exception as e:
        print(f"Error en procesamiento manual: {e}")
        return {"error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) > 1:
        ruta_excel = sys.argv[1]
    else:
        ruta_excel = '/Users/rodrigomunoz/Downloads/Lista de Trabajo - Por Estudio 2025-05-05 11-19-55.xlsx'
    
    analizar_excel_problematico(ruta_excel)