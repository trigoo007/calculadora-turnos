#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para procesar manualmente archivos Excel problemáticos
"""

import pandas as pd
import os
import sys
from openpyxl import load_workbook
import traceback

def procesar_excel_manualmente(ruta_excel, hoja_nombre=None):
    """
    Procesa manualmente un archivo Excel problemático, evitando conversiones automáticas
    que pueden causar errores.
    
    Args:
        ruta_excel: Ruta al archivo Excel
        hoja_nombre: Nombre de la hoja a procesar (si es None, se usa la primera hoja)
    
    Returns:
        Dictionary con información del procesamiento
    """
    try:
        print(f"Procesando archivo: {ruta_excel}")
        
        # Cargar el workbook
        wb = load_workbook(filename=ruta_excel, read_only=True, data_only=True)
        print(f"Hojas disponibles: {wb.sheetnames}")
        
        # Seleccionar la hoja
        if hoja_nombre and hoja_nombre in wb.sheetnames:
            sheet = wb[hoja_nombre]
        else:
            sheet = wb.active
            hoja_nombre = sheet.title
        
        print(f"Procesando hoja: {hoja_nombre}")
        
        # Procesar los datos manualmente
        data = []
        headers = []
        
        # Extraer encabezados (primera fila)
        row_count = 0
        header_row = None
        
        # Buscar encabezados (primera fila no vacía)
        for row in sheet.iter_rows():
            row_count += 1
            header_values = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
            
            # Verificar si la fila contiene datos reales
            if any(header_values) and header_row is None:
                header_row = row_count
                headers = header_values
                print(f"Encabezados encontrados en fila {header_row}: {headers}")
                break
        
        # Si no se encontraron encabezados, usar nombres de columna genéricos
        if not headers:
            max_col = sheet.max_column
            headers = [f"Columna{i}" for i in range(1, max_col + 1)]
            print(f"No se encontraron encabezados, usando genéricos: {headers}")
            header_row = 0
        
        # Procesar filas de datos
        data_row_count = 0
        
        for row in sheet.iter_rows(min_row=header_row + 1):
            row_data = []
            for cell in row:
                # Obtener el valor manteniendo el tipo original
                valor = cell.value
                
                # Manejo especial para valores problemáticos
                if isinstance(valor, str):
                    # Intentar limpiar valores como '0, 0'
                    if ',' in valor and valor.replace(',', '').replace(' ', '').isdigit():
                        valor = valor.replace(',', '.').replace(' ', '')
                
                row_data.append(valor)
            
            # Solo añadir filas que tengan al menos un valor no nulo
            if any(v is not None and str(v).strip() != "" for v in row_data):
                data.append(row_data)
                data_row_count += 1
        
        print(f"Procesadas {data_row_count} filas de datos")
        
        # Crear DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Limpiar las columnas a textos
        for col in df.columns:
            df[col] = df[col].astype(str)
            df[col] = df[col].replace('None', '')
            df[col] = df[col].str.strip()
        
        # Guardar como CSV para procesamiento posterior
        csv_path = os.path.join(os.path.dirname(ruta_excel), 
                              f"{os.path.splitext(os.path.basename(ruta_excel))[0]}_{hoja_nombre}_procesado.csv")
        
        df.to_csv(csv_path, index=False)
        print(f"Archivo CSV guardado en: {csv_path}")
        
        # Cerrar workbook
        wb.close()
        
        # Mostrar una muestra de los datos
        print("\nMuestra de datos (primeras 3 filas):")
        print(df.head(3))
        
        return {
            "exito": True,
            "csv_path": csv_path,
            "filas": len(df),
            "columnas": df.columns.tolist(),
            "hoja_procesada": hoja_nombre
        }
    
    except Exception as e:
        print(f"Error en procesamiento manual: {str(e)}")
        traceback.print_exc()
        return {"error": str(e)}

if __name__ == "__main__":
    if len(sys.argv) > 1:
        ruta_excel = sys.argv[1]
        hoja = sys.argv[2] if len(sys.argv) > 2 else None
    else:
        ruta_excel = '/Users/rodrigomunoz/Downloads/Lista de Trabajo - Por Estudio 2025-05-05 11-19-55.xlsx'
        hoja = None
    
    procesar_excel_manualmente(ruta_excel, hoja)